import os
import sys
import time
import asyncio
import csv
import json
import logging
from collections import deque
from typing import Optional, AsyncGenerator, Deque
from pathlib import Path

from fastmcp import FastMCP

# Set up logging
logger = logging.getLogger(__name__)

# Handle both package import and direct script execution
try:
    from .__version__ import __version__
except ImportError:
    # Add parent directory to path for direct script execution
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from server.__version__ import __version__

try:
    from pdfminer.high_level import extract_text as pdf_extract_text
except Exception:  # pragma: no cover
    pdf_extract_text = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

try:
    import docx
except Exception:  # pragma: no cover
    docx = None

try:
    import markdown
except Exception:  # pragma: no cover
    markdown = None

try:
    from markitdown import MarkItDown
except Exception:  # pragma: no cover
    MarkItDown = None

try:
    import fitz  # PyMuPDF for image extraction from PDFs
except Exception:  # pragma: no cover
    fitz = None


server = FastMCP("document-reader-mcp")


class SimpleRateLimiter:
    """In-memory sliding-window rate limiter (process-wide).

    Not suitable for multi-process or distributed deployments.
    """

    def __init__(self, max_calls: int, window_seconds: int = 60) -> None:
        self.max_calls = max_calls
        self.window_seconds = window_seconds
        self._timestamps: Deque[float] = deque()

    def allow(self) -> bool:
        now = time.time()
        window_start = now - self.window_seconds
        while self._timestamps and self._timestamps[0] < window_start:
            self._timestamps.popleft()
        if len(self._timestamps) >= self.max_calls:
            return False
        self._timestamps.append(now)
        return True


_rate_limit_per_minute_env = os.getenv("DOC_READER_RATE_LIMIT_PER_MINUTE", "60")
try:
    _rate_limit_per_minute = max(1, int(_rate_limit_per_minute_env))
except ValueError:
    _rate_limit_per_minute = 60

_rate_limiter = SimpleRateLimiter(max_calls=_rate_limit_per_minute, window_seconds=60)

# Maximum output text size in characters (to prevent context overflow)
_max_output_chars_env = os.getenv("DOC_READER_MAX_OUTPUT_CHARS", "100000")
try:
    _max_output_chars = max(1000, int(_max_output_chars_env))
except ValueError:
    _max_output_chars = 100000

# Default maximum rows for spreadsheets/CSV (0 means use max_output_chars limit only)
_default_max_rows_env = os.getenv("DOC_READER_DEFAULT_MAX_ROWS", "500")
try:
    _default_max_rows = max(0, int(_default_max_rows_env))
except ValueError:
    _default_max_rows = 500

# Default maximum pages for PDFs (0 means use max_output_chars limit only)
_default_max_pages_env = os.getenv("DOC_READER_DEFAULT_MAX_PAGES", "50")
try:
    _default_max_pages = max(0, int(_default_max_pages_env))
except ValueError:
    _default_max_pages = 50


def _enforce_rate_limit() -> None:
    if not _rate_limiter.allow():
        raise RuntimeError("Rate limit exceeded. Try again later or increase limits in configuration.")


def _truncate_output_if_needed(text: str, truncated_rows: bool = False, file_path: str = "") -> str:
    """Truncate output text if it exceeds maximum character limit and add warning."""
    if len(text) <= _max_output_chars:
        return text
    
    truncation_point = _max_output_chars
    # Try to truncate at a line boundary for cleaner output
    last_newline = text.rfind("\n", 0, _max_output_chars)
    if last_newline > _max_output_chars - 1000:  # Within reasonable distance
        truncation_point = last_newline
    
    truncated_text = text[:truncation_point]
    
    warning_msg = (
        f"\n\n[TRUNCATED: Output exceeded {_max_output_chars:,} character limit. "
        f"Original size: {len(text):,} characters. "
    )
    
    if file_path:
        warning_msg += f"File: {os.path.basename(file_path)}. "
    
    if truncated_rows:
        warning_msg += "Consider using max_rows or max_pages parameter to limit input. "
    
    warning_msg += (
        f"To increase limit, set DOC_READER_MAX_OUTPUT_CHARS environment variable.]"
    )
    
    return truncated_text + warning_msg


def _extract_text_from_pdf(path: str, max_pages: Optional[int] = None) -> str:
    if pdf_extract_text is None:
        raise RuntimeError(
            "pdfminer.six is not installed. To process PDF files, install it with: "
            "pip install pdfminer.six"
        )
    
    # Apply default page limit if none specified
    effective_max_pages = max_pages if max_pages is not None else _default_max_pages
    
    # Use pdfminer's maxpages to avoid parsing the whole file when limited
    maxpages_arg = 0 if effective_max_pages <= 0 else int(effective_max_pages)
    text = pdf_extract_text(path, maxpages=maxpages_arg)
    
    if effective_max_pages > 0:
        text += f"\n\n[INFO: Page limit of {effective_max_pages} applied. Use max_pages parameter to adjust.]"
    
    return _truncate_output_if_needed(text, truncated_rows=True, file_path=path)


def _extract_text_from_xlsx(path: str, max_rows: Optional[int] = None) -> str:
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is not installed. To process Excel files, install it with: "
            "pip install openpyxl"
        )
    
    # Apply default row limit if none specified
    effective_max_rows = max_rows if max_rows is not None else _default_max_rows
    
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        lines: list[str] = []
        rows_emitted = 0
        hit_row_limit = False
        
        for sheet in workbook.worksheets:
            lines.append(f"# Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = ["" if cell is None else str(cell) for cell in row]
                line = "\t".join(values).rstrip()
                if line:
                    lines.append(line)
                    rows_emitted += 1
                    if effective_max_rows > 0 and rows_emitted >= effective_max_rows:
                        hit_row_limit = True
                        break
            if hit_row_limit:
                break
            lines.append("")
        
        result = "\n".join(lines).strip()
        
        if hit_row_limit:
            result += f"\n\n[INFO: Row limit of {effective_max_rows} reached. Use max_rows parameter to adjust.]"
        
        return _truncate_output_if_needed(result, truncated_rows=True, file_path=path)
    finally:
        workbook.close()


def _extract_text_from_csv(path: str, max_rows: Optional[int] = None) -> str:
    """Extract text from CSV file using Python's built-in csv module."""
    # Apply default row limit if none specified
    effective_max_rows = max_rows if max_rows is not None else _default_max_rows
    
    lines: list[str] = []
    rows_read = 0
    hit_row_limit = False
    
    # Try different encodings to handle various CSV files
    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    
    for encoding in encodings:
        try:
            with open(path, 'r', encoding=encoding, newline='') as f:
                reader = csv.reader(f)
                for row in reader:
                    line = "\t".join(str(cell) for cell in row).rstrip()
                    if line:
                        lines.append(line)
                        rows_read += 1
                        if effective_max_rows > 0 and rows_read >= effective_max_rows:
                            hit_row_limit = True
                            break
            
            result = "\n".join(lines).strip()
            
            if hit_row_limit:
                result += f"\n\n[INFO: Row limit of {effective_max_rows} reached. Use max_rows parameter to adjust.]"
            
            return _truncate_output_if_needed(result, truncated_rows=True, file_path=path)
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception as e:
            if encoding == encodings[-1]:  # Last encoding attempt
                raise RuntimeError(f"Failed to read CSV file: {e}") from e
    
    if not lines:
        raise RuntimeError("Failed to decode CSV file with any supported encoding")
    
    result = "\n".join(lines).strip()
    if hit_row_limit:
        result += f"\n\n[INFO: Row limit of {effective_max_rows} reached. Use max_rows parameter to adjust.]"
    return _truncate_output_if_needed(result, truncated_rows=True, file_path=path)


def _extract_text_from_txt(path: str) -> str:
    """Extract text from plain text file."""
    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    
    for encoding in encodings:
        try:
            with open(path, 'r', encoding=encoding) as f:
                text = f.read()
                return _truncate_output_if_needed(text, truncated_rows=False, file_path=path)
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception as e:
            if encoding == encodings[-1]:
                raise RuntimeError(f"Failed to read text file: {e}") from e
    
    raise RuntimeError("Failed to decode text file with any supported encoding")


def _extract_text_from_json(path: str) -> str:
    """Extract text from JSON file (pretty-printed)."""
    encodings = ['utf-8', 'latin-1', 'cp1252']
    
    for encoding in encodings:
        try:
            with open(path, 'r', encoding=encoding) as f:
                data = json.load(f)
                text = json.dumps(data, indent=2, ensure_ascii=False)
                return _truncate_output_if_needed(text, truncated_rows=False, file_path=path)
        except (UnicodeDecodeError, UnicodeError):
            continue
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON file: {e}") from e
        except Exception as e:
            if encoding == encodings[-1]:
                raise RuntimeError(f"Failed to read JSON file: {e}") from e
    
    raise RuntimeError("Failed to decode JSON file with any supported encoding")


def _extract_text_from_markdown(path: str) -> str:
    """Extract text from Markdown file (as plain text or HTML)."""
    encodings = ['utf-8', 'latin-1', 'cp1252']
    
    for encoding in encodings:
        try:
            with open(path, 'r', encoding=encoding) as f:
                md_content = f.read()
                
            # If markdown library is available, optionally convert to HTML
            # For simplicity, return plain markdown text
            return _truncate_output_if_needed(md_content, truncated_rows=False, file_path=path)
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception as e:
            if encoding == encodings[-1]:
                raise RuntimeError(f"Failed to read Markdown file: {e}") from e
    
    raise RuntimeError("Failed to decode Markdown file with any supported encoding")


def _extract_text_from_docx(path: str) -> str:
    """Extract text from DOCX file."""
    if docx is None:
        raise RuntimeError(
            "python-docx is not installed. To process .docx files, install it with: "
            "pip install python-docx"
        )
    
    try:
        doc = docx.Document(path)
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        text = "\n".join(paragraphs)
        return _truncate_output_if_needed(text, truncated_rows=False, file_path=path)
    except Exception as e:
        raise RuntimeError(f"Failed to extract text from DOCX: {e}") from e


def _extract_images_from_pdf(pdf_path: str, output_dir: str, images_dirname: str) -> tuple[int, str, dict]:
    """
    Extract images from PDF using PyMuPDF and save them to output directory.
    
    Args:
        pdf_path: Path to the PDF file
        output_dir: Directory where images directory will be created
        images_dirname: Name of the images subdirectory
    
    Returns:
        Tuple of (image_count, images_dir_path, page_to_images_dict)
        page_to_images_dict maps page numbers to list of image filenames
    """
    if fitz is None:
        # If PyMuPDF is not available, return 0 images extracted
        return 0, "", {}
    
    try:
        images_dir = os.path.join(output_dir, images_dirname)
        image_count = 0
        page_to_images = {}
        
        # Open the PDF
        pdf_document = fitz.open(pdf_path)
        
        # Iterate through pages
        for page_num in range(len(pdf_document)):
            page = pdf_document[page_num]
            
            # Get list of images on the page
            image_list = page.get_images(full=True)
            
            if not image_list:
                continue
            
            page_images = []
            
            # Extract each image
            for img_index, img_info in enumerate(image_list):
                xref = img_info[0]  # XREF number of image
                
                try:
                    # Extract the image
                    base_image = pdf_document.extract_image(xref)
                    image_bytes = base_image["image"]
                    image_ext = base_image["ext"]  # png, jpeg, etc.
                    
                    # Create images directory if it doesn't exist
                    if image_count == 0:
                        os.makedirs(images_dir, exist_ok=True)
                    
                    # Save the image with a meaningful name
                    image_count += 1
                    img_filename = f"image_{image_count}.{image_ext}"
                    img_path = os.path.join(images_dir, img_filename)
                    
                    with open(img_path, "wb") as img_file:
                        img_file.write(image_bytes)
                    
                    page_images.append(img_filename)
                        
                except Exception as img_error:
                    # Skip images that fail to extract
                    logger.warning(f"Failed to extract image {img_index} from page {page_num + 1}: {img_error}")
                    continue
            
            if page_images:
                page_to_images[page_num + 1] = page_images
        
        pdf_document.close()
        
        logger.info(f"Successfully extracted {image_count} images from PDF")
        return image_count, images_dir if image_count > 0 else "", page_to_images
        
    except Exception as e:
        logger.error(f"Failed to extract images from PDF: {e}", exc_info=True)
        return 0, "", {}


@server.tool
def extract_text_from_file(
    path: str,
    max_pages: Optional[int] = None,
    max_rows: Optional[int] = None,
) -> str:
    """
    Extract plain text from local document files.

    Supported formats:
    - PDF (.pdf)
    - Excel (.xlsx, .xlsm, .xltx, .xltm)
    - Word (.docx)
    - CSV (.csv)
    - Plain text (.txt, .log, .text)
    - JSON (.json)
    - Markdown (.md, .markdown)

    Args:
        path: Absolute or relative file path on the local machine.
        max_pages: For PDFs, parse only the first N pages. If not specified, defaults to 50 pages.
            Set to 0 to disable page limit (not recommended for large files).
        max_rows: For spreadsheets and CSV, parse only N data rows across all sheets. 
            If not specified, defaults to 500 rows. Set to 0 to disable row limit 
            (not recommended for large files).

    Returns:
        Extracted plain text as a string. Output is automatically truncated at 100,000 
        characters by default to prevent context overflow. Set DOC_READER_MAX_OUTPUT_CHARS 
        environment variable to adjust this limit.
    """
    _enforce_rate_limit()
    if not path or not isinstance(path, str):
        raise ValueError("path must be a non-empty string")

    expanded_path = os.path.expanduser(path)
    if not os.path.isfile(expanded_path):
        raise FileNotFoundError(f"File not found: {expanded_path}")

    file_size = os.path.getsize(expanded_path)
    if file_size > 100 * 1024 * 1024:
        raise ValueError("File too large; limit is 100MB")

    _, ext = os.path.splitext(expanded_path)
    ext_lower = ext.lower()

    # Route to appropriate extractor based on file extension
    if ext_lower == ".pdf":
        text = _extract_text_from_pdf(expanded_path, max_pages=max_pages)
    elif ext_lower in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        text = _extract_text_from_xlsx(expanded_path, max_rows=max_rows)
    elif ext_lower == ".csv":
        text = _extract_text_from_csv(expanded_path, max_rows=max_rows)
    elif ext_lower in (".txt", ".log", ".text"):
        text = _extract_text_from_txt(expanded_path)
    elif ext_lower == ".json":
        text = _extract_text_from_json(expanded_path)
    elif ext_lower in (".md", ".markdown"):
        text = _extract_text_from_markdown(expanded_path)
    elif ext_lower == ".docx":
        text = _extract_text_from_docx(expanded_path)
    else:
        raise ValueError(
            f"Unsupported file type: {ext_lower}. "
            f"Supported: .pdf, .xlsx, .csv, .txt, .json, .md, .docx"
        )

    return text


@server.tool
async def extract_text_from_file_stream(
    path: str,
    max_pages: Optional[int] = None,
    max_rows: Optional[int] = None,
    chunk_size: int = 4096,
) -> AsyncGenerator[str, None]:
    """
    Stream plain text chunks from local document files.

    Supported formats: PDF, Excel, CSV, TXT, JSON, Markdown, DOCX

    Args:
        path: Absolute or relative file path on the local machine.
        max_pages: For PDFs, parse only the first N pages. If not specified, defaults to 50 pages.
            Set to 0 to disable page limit (not recommended for large files).
        max_rows: For spreadsheets and CSV, parse only N data rows across all sheets. 
            If not specified, defaults to 500 rows. Set to 0 to disable row limit 
            (not recommended for large files).
        chunk_size: Approximate maximum characters per streamed chunk. Actual chunk sizes may
            vary slightly.

    Yields:
        Text chunks as strings until the entire document (or capped portion) has been sent.
        Streaming respects the same character limits as non-streaming extraction.
    """
    _enforce_rate_limit()
    if not path or not isinstance(path, str):
        raise ValueError("path must be a non-empty string")

    expanded_path = os.path.expanduser(path)
    if not os.path.isfile(expanded_path):
        raise FileNotFoundError(f"File not found: {expanded_path}")

    file_size = os.path.getsize(expanded_path)
    if file_size > 100 * 1024 * 1024:
        raise ValueError("File too large; limit is 100MB")

    _, ext = os.path.splitext(expanded_path)
    ext_lower = ext.lower()
    chunk_size = max(512, int(chunk_size))

    if ext_lower == ".pdf":
        # Extract text upfront with page cap, then stream in fixed-size chunks
        text = _extract_text_from_pdf(expanded_path, max_pages=max_pages)
        for i in range(0, len(text), chunk_size):
            yield text[i : i + chunk_size]
            await asyncio.sleep(0)
            
    elif ext_lower in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        # Stream rows as they are read, accumulating into approx chunk_size blocks
        if load_workbook is None:
            raise RuntimeError(
                "openpyxl is not installed. To process Excel files, install it with: "
                "pip install openpyxl"
            )
        
        # Apply default row limit if none specified
        effective_max_rows = max_rows if max_rows is not None else _default_max_rows
        
        workbook = load_workbook(filename=expanded_path, data_only=True, read_only=True)
        try:
            buffer_lines: list[str] = []
            buffer_len = 0
            rows_emitted = 0
            total_chars_emitted = 0
            hit_row_limit = False
            
            for sheet in workbook.worksheets:
                header = f"# Sheet: {sheet.title}"
                if buffer_len + len(header) + 1 > chunk_size and buffer_lines:
                    chunk_text = "\n".join(buffer_lines)
                    total_chars_emitted += len(chunk_text)
                    if total_chars_emitted > _max_output_chars:
                        break
                    yield chunk_text
                    buffer_lines = []
                    buffer_len = 0
                    await asyncio.sleep(0)
                buffer_lines.append(header)
                buffer_len += len(header) + 1

                for row in sheet.iter_rows(values_only=True):
                    values = ["" if cell is None else str(cell) for cell in row]
                    line = "\t".join(values).rstrip()
                    if not line:
                        continue
                    if buffer_len + len(line) + 1 > chunk_size and buffer_lines:
                        chunk_text = "\n".join(buffer_lines)
                        total_chars_emitted += len(chunk_text)
                        if total_chars_emitted > _max_output_chars:
                            break
                        yield chunk_text
                        buffer_lines = []
                        buffer_len = 0
                        await asyncio.sleep(0)
                    buffer_lines.append(line)
                    buffer_len += len(line) + 1
                    rows_emitted += 1
                    if effective_max_rows > 0 and rows_emitted >= effective_max_rows:
                        hit_row_limit = True
                        break
                if hit_row_limit or total_chars_emitted > _max_output_chars:
                    break
                    
            if buffer_lines:
                chunk_text = "\n".join(buffer_lines)
                total_chars_emitted += len(chunk_text)
                yield chunk_text
                
            # Send info message if limits were hit
            if hit_row_limit:
                yield f"\n\n[INFO: Row limit of {effective_max_rows} reached. Use max_rows parameter to adjust.]"
            if total_chars_emitted > _max_output_chars:
                yield f"\n\n[TRUNCATED: Output exceeded {_max_output_chars:,} character limit.]"
        finally:
            workbook.close()
            
    elif ext_lower == ".csv":
        # Stream CSV rows with buffering
        # Apply default row limit if none specified
        effective_max_rows = max_rows if max_rows is not None else _default_max_rows
        
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        success = False
        
        for encoding in encodings:
            try:
                buffer_lines: list[str] = []
                buffer_len = 0
                rows_emitted = 0
                total_chars_emitted = 0
                hit_row_limit = False
                
                with open(expanded_path, 'r', encoding=encoding, newline='') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        line = "\t".join(str(cell) for cell in row).rstrip()
                        if not line:
                            continue
                        if buffer_len + len(line) + 1 > chunk_size and buffer_lines:
                            chunk_text = "\n".join(buffer_lines)
                            total_chars_emitted += len(chunk_text)
                            if total_chars_emitted > _max_output_chars:
                                break
                            yield chunk_text
                            buffer_lines = []
                            buffer_len = 0
                            await asyncio.sleep(0)
                        buffer_lines.append(line)
                        buffer_len += len(line) + 1
                        rows_emitted += 1
                        if effective_max_rows > 0 and rows_emitted >= effective_max_rows:
                            hit_row_limit = True
                            break
                    
                    if buffer_lines:
                        chunk_text = "\n".join(buffer_lines)
                        total_chars_emitted += len(chunk_text)
                        yield chunk_text
                        
                    # Send info message if limits were hit
                    if hit_row_limit:
                        yield f"\n\n[INFO: Row limit of {effective_max_rows} reached. Use max_rows parameter to adjust.]"
                    if total_chars_emitted > _max_output_chars:
                        yield f"\n\n[TRUNCATED: Output exceeded {_max_output_chars:,} character limit.]"
                        
                success = True
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
                
        if not success:
            raise RuntimeError("Failed to decode CSV file with any supported encoding")
            
    elif ext_lower in (".txt", ".log", ".text", ".md", ".markdown"):
        # Stream text files in chunks
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        success = False
        
        for encoding in encodings:
            try:
                total_chars_emitted = 0
                with open(expanded_path, 'r', encoding=encoding) as f:
                    while True:
                        chunk = f.read(chunk_size)
                        if not chunk:
                            break
                        total_chars_emitted += len(chunk)
                        if total_chars_emitted > _max_output_chars:
                            # Truncate the final chunk
                            chars_over = total_chars_emitted - _max_output_chars
                            chunk = chunk[:-chars_over] if chars_over < len(chunk) else ""
                            if chunk:
                                yield chunk
                            yield f"\n\n[TRUNCATED: Output exceeded {_max_output_chars:,} character limit.]"
                            break
                        yield chunk
                        await asyncio.sleep(0)
                success = True
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
                
        if not success:
            raise RuntimeError("Failed to decode text file with any supported encoding")
            
    elif ext_lower == ".json":
        # For JSON, extract all then stream in chunks (can't partially parse JSON)
        text = _extract_text_from_json(expanded_path)
        for i in range(0, len(text), chunk_size):
            yield text[i : i + chunk_size]
            await asyncio.sleep(0)
            
    elif ext_lower == ".docx":
        # For DOCX, extract all then stream in chunks
        text = _extract_text_from_docx(expanded_path)
        for i in range(0, len(text), chunk_size):
            yield text[i : i + chunk_size]
            await asyncio.sleep(0)
            
    else:
        raise ValueError(
            f"Unsupported file type: {ext_lower}. "
            f"Supported: .pdf, .xlsx, .csv, .txt, .json, .md, .docx"
        )


@server.tool
def convert_to_markdown(
    path: str,
    output_dir: Optional[str] = None,
    output_filename: Optional[str] = None,
) -> dict:
    """
    Convert various document formats to Markdown, extracting images when applicable.
    
    **Important**: This tool converts the ENTIRE document and saves it to a file.
    It ignores the DOC_READER_DEFAULT_MAX_ROWS, DOC_READER_DEFAULT_MAX_PAGES, and 
    DOC_READER_MAX_OUTPUT_CHARS environment variables. Only the preview returned to 
    the AI is limited to protect context - the saved file contains the complete document.
    
    Supported formats:
    - PDF (.pdf) - with image extraction
    - Excel (.xlsx, .xlsm, .xltx, .xltm) - converted to markdown tables
    - Word (.docx) - with image extraction
    - CSV (.csv) - converted to markdown tables
    - PowerPoint (.pptx) - text and images
    - HTML (.html, .htm)
    - Plain text (.txt, .log)
    - Images (.jpg, .jpeg, .png) - with OCR if available
    
    Args:
        path: Absolute or relative path to the file to convert.
        output_dir: Directory where the markdown file and images will be saved.
            If not specified, saves in the same directory as the source file.
        output_filename: Name for the output markdown file (without extension).
            If not specified, uses the source filename with .md extension.
    
    Returns:
        Dictionary containing:
        - markdown_path: Path to the saved markdown file (contains FULL content, not truncated)
        - images_dir: Path to the directory containing extracted images (if any)
        - image_count: Number of images extracted
        - markdown_preview: First 500 characters preview (truncated for AI context protection)
        - file_size_chars: Total character count of the saved markdown file
    """
    _enforce_rate_limit()
    
    if MarkItDown is None:
        raise RuntimeError(
            "markitdown is not installed. To use conversion features, install it with: "
            "pip install markitdown"
        )
    
    if not path or not isinstance(path, str):
        raise ValueError("path must be a non-empty string")
    
    expanded_path = os.path.expanduser(path)
    if not os.path.isfile(expanded_path):
        raise FileNotFoundError(f"File not found: {expanded_path}")
    
    file_size = os.path.getsize(expanded_path)
    if file_size > 100 * 1024 * 1024:
        raise ValueError("File too large; limit is 100MB")
    
    # Determine output directory
    if output_dir:
        output_directory = os.path.expanduser(output_dir)
    else:
        output_directory = os.path.dirname(expanded_path)
    
    # Create output directory if it doesn't exist
    os.makedirs(output_directory, exist_ok=True)
    
    # Determine output filename
    source_basename = os.path.basename(expanded_path)
    source_name, _ = os.path.splitext(source_basename)
    
    if output_filename:
        md_filename = output_filename if output_filename.endswith('.md') else f"{output_filename}.md"
    else:
        md_filename = f"{source_name}.md"
    
    md_path = os.path.join(output_directory, md_filename)
    
    # Create images directory for this document
    images_dirname = f"{source_name}_images"
    images_dir = os.path.join(output_directory, images_dirname)
    
    try:
        # Initialize MarkItDown converter
        converter = MarkItDown()
        
        # Convert the document
        result = converter.convert(expanded_path)
        markdown_content = result.text_content
        
        # Check file extension to determine if we need special image handling
        _, ext = os.path.splitext(expanded_path)
        ext_lower = ext.lower()
        
        # Check if there are embedded images in the result
        image_count = 0
        if hasattr(result, 'images') and result.images:
            os.makedirs(images_dir, exist_ok=True)
            
            # Save images and update references in markdown
            for idx, img_data in enumerate(result.images):
                img_filename = f"image_{idx + 1}.png"
                img_path = os.path.join(images_dir, img_filename)
                
                # Save the image
                with open(img_path, 'wb') as f:
                    f.write(img_data)
                
                image_count += 1
                
                # Update markdown to reference the saved image
                relative_img_path = f"{images_dirname}/{img_filename}"
                # This is a simplified approach; actual implementation may vary
                # based on how markitdown structures image references
        
        # If markitdown doesn't provide direct image access, check for image links in markdown
        # and extract them if they're data URIs
        if 'data:image' in markdown_content and image_count == 0:
            import re
            import base64
            
            # Find all data URI images
            data_uri_pattern = r'!\[([^\]]*)\]\(data:image/([^;]+);base64,([^)]+)\)'
            matches = re.findall(data_uri_pattern, markdown_content)
            
            if matches:
                os.makedirs(images_dir, exist_ok=True)
                
                for idx, (alt_text, img_format, base64_data) in enumerate(matches):
                    img_filename = f"image_{idx + 1}.{img_format}"
                    img_path = os.path.join(images_dir, img_filename)
                    
                    # Decode and save the image
                    try:
                        img_bytes = base64.b64decode(base64_data)
                        with open(img_path, 'wb') as f:
                            f.write(img_bytes)
                        
                        image_count += 1
                        
                        # Replace data URI with file reference
                        relative_img_path = f"{images_dirname}/{img_filename}"
                        old_ref = f'![{alt_text}](data:image/{img_format};base64,{base64_data})'
                        new_ref = f'![{alt_text}]({relative_img_path})'
                        markdown_content = markdown_content.replace(old_ref, new_ref)
                    except Exception as e:
                        # If image extraction fails, continue with next image
                        logger.warning(f"Failed to extract data URI image {idx + 1}: {e}")
        
        # For PDFs, use PyMuPDF to extract images if markitdown didn't extract them
        if ext_lower == ".pdf" and image_count == 0:
            logger.info(f"Attempting to extract images from PDF using PyMuPDF: {expanded_path}")
            extracted_count, extracted_dir, page_to_images = _extract_images_from_pdf(
                expanded_path, output_directory, images_dirname
            )
            logger.info(f"PDF image extraction completed: {extracted_count} images extracted")
            
            if extracted_count > 0:
                image_count = extracted_count
                images_dir = extracted_dir
                
                # Insert images throughout the document based on page breaks
                # Look for form feed characters (\f) which markitdown uses as page markers
                if '\f' in markdown_content and page_to_images:
                    # Split content by page breaks
                    pages = markdown_content.split('\f')
                    reconstructed_content = []
                    
                    for page_num, page_content in enumerate(pages, start=1):
                        reconstructed_content.append(page_content)
                        
                        # If this page has images, insert them
                        if page_num in page_to_images:
                            reconstructed_content.append("\n\n")
                            for img_filename in page_to_images[page_num]:
                                relative_img_path = f"{images_dirname}/{img_filename}"
                                reconstructed_content.append(f"![Image from page {page_num}]({relative_img_path})\n\n")
                        
                        # Add page break back (except for last page)
                        if page_num < len(pages):
                            reconstructed_content.append('\f')
                    
                    markdown_content = ''.join(reconstructed_content)
                else:
                    # Fallback: add images at the end if no page breaks found
                    markdown_content += "\n\n## Extracted Images\n\n"
                    markdown_content += "*The following images were extracted from the PDF:*\n\n"
                    
                    for page_num in sorted(page_to_images.keys()):
                        markdown_content += f"\n### Images from Page {page_num}\n\n"
                        for img_filename in page_to_images[page_num]:
                            relative_img_path = f"{images_dirname}/{img_filename}"
                            markdown_content += f"![Image]({relative_img_path})\n\n"
        
        # Save the FULL markdown content to file (no truncation)
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)
        
        # Prepare preview for return value (truncated for AI context)
        # Only return first 500 chars as preview, plus info about size
        original_length = len(markdown_content)
        preview = markdown_content[:500]
        if original_length > 500:
            preview += f"\n\n... (truncated preview, full file has {original_length:,} characters)"
        
        return {
            "markdown_path": md_path,
            "images_dir": images_dir if image_count > 0 else None,
            "image_count": image_count,
            "markdown_preview": preview,
            "file_size_chars": original_length,
            "status": "success",
            "message": f"Successfully converted {source_basename} to Markdown ({original_length:,} characters)"
        }
        
    except Exception as e:
        raise RuntimeError(f"Failed to convert document to Markdown: {e}") from e


if __name__ == "__main__":
    server.run()
